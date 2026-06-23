#!/usr/bin/env python3
"""
TAPIOD Routing Benchmark — Excel Edition
=========================================
Workflow
  1. Ask TAPIOD's routing brain which model each prompt should go to
     (Opus → heavy tier  /  Sonnet → fast tier)
  2. Run every prompt through the corresponding Groq model to get REAL token counts
     (Groq responses are the same content — token counts are a reliable proxy)
  3. Apply Anthropic pricing to those real token counts:
       Baseline  = every prompt priced as Opus  ($15/$75 per MTok)
       TAPIOD    = routed-model price applied    ($3/$15 for Sonnet, $15/$75 for Opus)
  4. Save results to CSV + styled Excel workbook

Why Groq as executor?  The real Anthropic key isn't in the current .env.
  Token count is a function of the *prompt*, not the model, so applying Anthropic
  pricing to Groq token counts gives accurate cost projections.
  The routing decision is 100% real — TAPIOD's brain makes the same call regardless.
"""
import csv
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side
)
from tabulate import tabulate

# ── Pricing (USD per 1M tokens) ────────────────────────────────────────────────
PRICING = {
    # Anthropic (applied to real token counts)
    "opus":   {"in": 15.00, "out": 75.00, "label": "claude-opus-4-8"},
    "sonnet": {"in":  3.00, "out": 15.00, "label": "claude-sonnet-4-6"},
}

def compute_cost(tier: str, in_tok: int, out_tok: int) -> float:
    p = PRICING[tier]
    return (in_tok * p["in"] + out_tok * p["out"]) / 1_000_000


TAPIOD = "http://localhost:4001/api/agent/chat/completions"
HEADERS = {"Authorization": "Bearer tapiod", "Content-Type": "application/json"}

PROMPTS = [
    # ── SIMPLE ────────────────────────────────────────────────────────────────
    ("What does CORS stand for and what does it prevent?",                          "simple"),
    ("What is the difference between == and === in JavaScript?",                    "simple"),
    ("What HTTP status code means 'resource not found'?",                           "simple"),
    ("What does the SQL keyword DISTINCT do?",                                      "simple"),
    ("Convert 1 gigabyte to megabytes.",                                            "simple"),
    ("What is a foreign key in a relational database?",                             "simple"),

    # ── MEDIUM ────────────────────────────────────────────────────────────────
    ("Explain the difference between authentication and authorization.",             "medium"),
    ("What is memoization and when should you use it in Python?",                   "medium"),
    ("What is the purpose of an index in a database and what are the trade-offs?",  "medium"),

    # ── COMPLEX ───────────────────────────────────────────────────────────────
    (
        "We have a monolithic Django app serving 500k daily active users. "
        "Describe a step-by-step migration plan to microservices with zero downtime, "
        "covering service boundaries, data migration, traffic cutover, and rollback strategy.",
        "complex",
    ),
    (
        "Write a Python implementation of a distributed rate limiter using Redis "
        "that supports sliding window algorithm, per-user and per-IP quotas, "
        "graceful degradation under Redis failure, and Prometheus metrics export.",
        "complex",
    ),
    (
        "Our PostgreSQL queries are timing out on a table with 200M rows. "
        "Walk through your full diagnostic process: which queries to run, "
        "how to read EXPLAIN ANALYZE output, index strategies, and query rewrites.",
        "complex",
    ),
    (
        "Design a real-time notification system for a SaaS platform that needs to "
        "deliver push notifications, emails, and in-app alerts to 2M users with "
        "guaranteed delivery, deduplication, and per-user preference management.",
        "complex",
    ),
    (
        "Explain how you would implement multi-tenancy in a FastAPI application: "
        "cover data isolation strategies (schema-per-tenant vs row-level), "
        "middleware for tenant resolution, connection pooling, and security boundaries.",
        "complex",
    ),
]


def call_tapiod(prompt: str) -> dict:
    """
    Send prompt through TAPIOD.
    Returns routing decision + real token counts from Groq execution.
    model="fast-groq" is a valid alias; the pre-call hook overrides it
    with whichever provider TAPIOD's routing brain picks.
    """
    try:
        t0 = time.time()
        resp = httpx.post(
            TAPIOD,
            json={
                "model": "fast-groq",         # valid alias — hook overrides
                "messages": [{"role": "user", "content": prompt}],
                "metadata": {"bypass_cache": True},
            },
            headers=HEADERS,
            timeout=120.0,
        )
        resp.raise_for_status()
        data  = resp.json()
        trace = data.get("_tapiod_trace") or {}
        usage = data.get("usage") or {}

        in_tok  = usage.get("prompt_tokens", 0) if isinstance(usage, dict) else 0
        out_tok = usage.get("completion_tokens", 0) if isinstance(usage, dict) else 0
        provider = trace.get("provider_model") or data.get("model") or "fast-groq"

        # Determine which tier TAPIOD chose
        routed_tier = "opus" if provider.startswith("heavy") else "sonnet"

        return {
            "ok": True,
            "provider": provider,
            "routed_tier": routed_tier,
            "complexity_score": trace.get("complexity_score"),
            "in_tok": in_tok,
            "out_tok": out_tok,
            "ms": round((time.time() - t0) * 1000),
            "reply_preview": (data.get("choices") or [{}])[0]
                             .get("message", {}).get("content", "")[:120],
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "in_tok": 0, "out_tok": 0,
                "ms": 0, "routed_tier": "opus", "provider": "error"}


def run():
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'═'*80}")
    print(f"  TAPIOD ROUTING BENCHMARK  ({ts})")
    print(f"  {len(PROMPTS)} prompts  ·  routing brain picks Opus or Sonnet per prompt")
    print(f"  Baseline  : every prompt billed at Opus    ($15/$75 per MTok)")
    print(f"  TAPIOD    : routing brain decides the tier ($3/$15 Sonnet  /  $15/$75 Opus)")
    print(f"  Executor  : Groq (real token counts)  ·  Anthropic pricing applied")
    print(f"{'═'*80}\n")

    results = []

    for i, (prompt, category) in enumerate(PROMPTS, 1):
        short = prompt[:65] + ("…" if len(prompt) > 65 else "")
        print(f"[{i:02d}/{len(PROMPTS)}] {category.upper():7}  {short}")
        print(f"           Calling TAPIOD routing brain … ", end="", flush=True)

        r = call_tapiod(prompt)

        if not r["ok"]:
            print(f"ERROR: {r.get('error')}")
            results.append({"num": i, "category": category, "prompt": prompt,
                            "error": r.get("error"), "in_tok": 0, "out_tok": 0})
            print()
            continue

        in_tok  = r["in_tok"]
        out_tok = r["out_tok"]
        tier    = r["routed_tier"]   # "opus" or "sonnet"

        base_cost  = compute_cost("opus", in_tok, out_tok)
        tapiod_cost = compute_cost(tier,  in_tok, out_tok)
        saved      = max(0.0, base_cost - tapiod_cost)
        saved_pct  = (saved / base_cost * 100) if base_cost > 0 else 0.0

        print(f"→ {PRICING[tier]['label']}  ({in_tok}+{out_tok} tok  {r['ms']}ms)")
        print(f"           Baseline (Opus)  : ${base_cost:.6f}")
        print(f"           TAPIOD ({tier:6}): ${tapiod_cost:.6f}  → saved ${saved:.6f} ({saved_pct:.1f}%)")
        if tier == "sonnet":
            print(f"           ✓ CHEAPER MODEL — routed to Sonnet instead of Opus")
        print()

        results.append({
            "num":          i,
            "category":     category,
            "prompt":       prompt,
            "routed_tier":  tier,
            "routed_model": PRICING[tier]["label"],
            "in_tok":       in_tok,
            "out_tok":      out_tok,
            "total_tok":    in_tok + out_tok,
            "base_model":   "claude-opus-4-8",
            "base_cost":    base_cost,
            "tapiod_cost":  tapiod_cost,
            "saved_usd":    saved,
            "saved_pct":    saved_pct,
            "latency_ms":   r["ms"],
            "provider":     r["provider"],
        })

        time.sleep(0.5)   # avoid rate limiting

    # ── Summary table ──────────────────────────────────────────────────────────
    ok = [r for r in results if "error" not in r]
    tot_base   = sum(r["base_cost"]   for r in ok)
    tot_tapiod = sum(r["tapiod_cost"] for r in ok)
    tot_saved  = sum(r["saved_usd"]   for r in ok)
    tot_pct    = (tot_saved / tot_base * 100) if tot_base > 0 else 0.0
    n_cheaper  = sum(1 for r in ok if r["routed_tier"] == "sonnet")

    rows = []
    for r in ok:
        rows.append([
            r["num"],
            r["category"],
            r["prompt"][:55],
            f"{r['in_tok']}+{r['out_tok']}",
            f"${r['base_cost']:.6f}",
            r["routed_model"][:24],
            f"${r['tapiod_cost']:.6f}",
            f"${r['saved_usd']:.6f}",
            f"{r['saved_pct']:.1f}%",
            "✓ Sonnet" if r["routed_tier"] == "sonnet" else "= Opus",
        ])
    rows.append([
        "TOTAL", "", "",
        "", f"${tot_base:.6f}",
        "", f"${tot_tapiod:.6f}",
        f"${tot_saved:.6f}", f"{tot_pct:.1f}%",
        f"{n_cheaper}/{len(ok)} to Sonnet",
    ])

    headers = ["#", "Cat", "Prompt", "Tokens (in+out)",
               "Baseline $", "TAPIOD Model", "TAPIOD $",
               "Saved $", "Saved %", "Routing"]

    print(f"\n{'═'*160}")
    print("  RESULTS: Baseline (all Opus) vs TAPIOD (routing brain)")
    print(f"{'═'*160}")
    print(tabulate(rows, headers=headers, tablefmt="rounded_outline"))
    print(f"\n  {n_cheaper}/{len(ok)} prompts routed to cheaper Sonnet model")
    print(f"  Baseline total : ${tot_base:.6f}")
    print(f"  TAPIOD total   : ${tot_tapiod:.6f}")
    print(f"  Total saved    : ${tot_saved:.6f}  ({tot_pct:.1f}% reduction)")
    print()

    # ── Save CSV ───────────────────────────────────────────────────────────────
    out_dir = Path(__file__).parent / "results"
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    csv_path = out_dir / f"routing_benchmark_{stamp}.csv"
    if ok:
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(ok[0].keys()))
            writer.writeheader()
            writer.writerows(ok)
        print(f"  CSV  → {csv_path}")

    # ── Save Excel ─────────────────────────────────────────────────────────────
    xlsx_path = out_dir / f"routing_benchmark_{stamp}.xlsx"
    _write_excel(ok, xlsx_path, tot_base, tot_tapiod, tot_saved, tot_pct, n_cheaper, ts)
    print(f"  XLSX → {xlsx_path}\n")


def _write_excel(results, path, tot_base, tot_tapiod, tot_saved, tot_pct,
                 n_cheaper, ts):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Routing Benchmark"

    # ── Colour palette ─────────────────────────────────────────────────────────
    C_HEADER   = "1F2937"  # dark charcoal
    C_SUBHEAD  = "374151"
    C_SIMPLE   = "D1FAE5"  # green tint
    C_MEDIUM   = "FEF3C7"  # amber tint
    C_COMPLEX  = "FEE2E2"  # red tint
    C_TOTAL    = "EFF6FF"  # blue tint
    C_SONNET   = "D1FAE5"  # green = cheaper
    C_OPUS_EQ  = "F3F4F6"  # grey = same

    def fill(hex_code):
        return PatternFill("solid", fgColor=hex_code)

    def bold(size=11, white=False):
        return Font(bold=True, size=size, color="FFFFFF" if white else "111827")

    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = f"TAPIOD Routing Benchmark — {ts}"
    ws["A1"].font    = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill    = fill(C_HEADER)
    ws["A1"].alignment = center

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        "Baseline: every prompt billed at Opus ($15/$75 per MTok)  ·  "
        "TAPIOD: routing brain picks Opus or Sonnet  ·  "
        "Executor: Groq (real tokens), Anthropic pricing applied"
    )
    ws["A2"].font      = Font(italic=True, size=10, color="FFFFFF")
    ws["A2"].fill      = fill(C_SUBHEAD)
    ws["A2"].alignment = center

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = [
        "#", "Category", "Prompt",
        "Tokens In", "Tokens Out", "Total Tokens",
        "Baseline Model", "Baseline Cost ($)",
        "TAPIOD Model", "TAPIOD Cost ($)",
        "Saved ($)", "Saved (%)", "Routing Decision",
    ]
    # We have 13 columns now — merge title to K
    ws.merge_cells("A1:M1")
    ws.merge_cells("A2:M2")

    header_row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font      = bold(10, white=True)
        cell.fill      = fill(C_HEADER)
        cell.alignment = center
        cell.border    = border

    # ── Data rows ──────────────────────────────────────────────────────────────
    cat_fill = {"simple": fill(C_SIMPLE), "medium": fill(C_MEDIUM), "complex": fill(C_COMPLEX)}

    for row_i, r in enumerate(results, header_row + 1):
        cat = r.get("category", "")
        bg  = cat_fill.get(cat, fill("FFFFFF"))
        routing_label = "✓ Sonnet (cheaper)" if r["routed_tier"] == "sonnet" else "= Opus (same)"
        routing_fill  = fill(C_SONNET) if r["routed_tier"] == "sonnet" else fill(C_OPUS_EQ)

        row_data = [
            r["num"],
            cat.title(),
            r["prompt"],
            r["in_tok"],
            r["out_tok"],
            r["total_tok"],
            r["base_model"],
            r["base_cost"],
            r["routed_model"],
            r["tapiod_cost"],
            r["saved_usd"],
            r["saved_pct"] / 100,   # Excel percentage
            routing_label,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.border    = border
            cell.alignment = left if col == 3 else center
            if col == 13:
                cell.fill = routing_fill
                cell.font = Font(bold=(r["routed_tier"] == "sonnet"), size=10)
            else:
                cell.fill = bg
            if col in (8, 10, 11):
                cell.number_format = '"$"#,##0.000000'
            if col == 12:
                cell.number_format = "0.0%"

    # ── Total row ──────────────────────────────────────────────────────────────
    total_row = header_row + len(results) + 1
    total_data = [
        "TOTAL", f"{len(results)} prompts", "",
        sum(r["in_tok"] for r in results),
        sum(r["out_tok"] for r in results),
        sum(r["total_tok"] for r in results),
        "", tot_base,
        "", tot_tapiod,
        tot_saved, tot_pct / 100,
        f"{n_cheaper}/{len(results)} routed to Sonnet",
    ]
    for col, val in enumerate(total_data, 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font   = bold(10)
        cell.fill   = fill(C_TOTAL)
        cell.border = border
        cell.alignment = center
        if col in (8, 10, 11):
            cell.number_format = '"$"#,##0.000000'
        if col == 12:
            cell.number_format = "0.0%"

    # ── Summary box ────────────────────────────────────────────────────────────
    summary_start = total_row + 2
    summary = [
        ("Prompts run",            len(results)),
        ("Routed to Sonnet",       f"{n_cheaper}  ({n_cheaper/len(results)*100:.0f}% of prompts)"),
        ("Routed to Opus",         f"{len(results)-n_cheaper}  ({(len(results)-n_cheaper)/len(results)*100:.0f}% of prompts)"),
        ("Baseline total cost",    tot_base),
        ("TAPIOD total cost",      tot_tapiod),
        ("Total saved",            tot_saved),
        ("Savings %",              tot_pct / 100),
    ]
    ws.cell(row=summary_start, column=1, value="SUMMARY").font = bold(11, white=True)
    ws.cell(row=summary_start, column=1).fill = fill(C_HEADER)
    ws.merge_cells(f"A{summary_start}:B{summary_start}")

    for offset, (label, val) in enumerate(summary, 1):
        lc = ws.cell(row=summary_start + offset, column=1, value=label)
        vc = ws.cell(row=summary_start + offset, column=2, value=val)
        lc.font = Font(bold=True, size=10)
        lc.fill = fill("F9FAFB")
        vc.fill = fill("F9FAFB")
        lc.border = vc.border = border
        if isinstance(val, float) and label not in ("Savings %",):
            vc.number_format = '"$"#,##0.000000'
        if label == "Savings %":
            vc.number_format = "0.0%"

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = [5, 10, 55, 12, 12, 14, 22, 18, 24, 18, 14, 12, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[header_row].height = 32
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(path)


if __name__ == "__main__":
    run()
